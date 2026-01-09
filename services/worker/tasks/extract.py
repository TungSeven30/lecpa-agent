"""Text extraction tasks for different document types."""

from dataclasses import dataclass
from pathlib import Path

import fitz  # pymupdf
import structlog
from docx import Document as DocxDocument
from openpyxl import load_workbook

from main import app

logger = structlog.get_logger()


@dataclass
class ExtractionResult:
    """Result from text extraction."""

    text: str
    page_count: int
    page_texts: list[str]  # Text per page
    metadata: dict


@app.task(bind=True, name="tasks.extract.extract_pdf")
def extract_pdf(self, file_path: str) -> dict:
    """Extract text from a PDF file.

    Args:
        file_path: Path to the PDF file

    Returns:
        Dict with extracted text, page count, and page texts
    """
    logger.info("Extracting text from PDF", file_path=file_path)

    try:
        doc = fitz.open(file_path)
        page_texts = []
        full_text = []

        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            page_texts.append(text)
            full_text.append(f"[PAGE {page_num + 1}]\n{text}")

        metadata = doc.metadata or {}
        page_count = len(doc)
        doc.close()

        return {
            "text": "\n\n".join(full_text),
            "page_count": page_count,
            "page_texts": page_texts,
            "metadata": {
                "title": metadata.get("title", ""),
                "author": metadata.get("author", ""),
                "subject": metadata.get("subject", ""),
            },
        }

    except Exception as e:
        logger.error("PDF extraction failed", error=str(e), file_path=file_path)
        raise


@app.task(bind=True, name="tasks.extract.extract_docx")
def extract_docx(self, file_path: str) -> dict:
    """Extract text from a DOCX file.

    Args:
        file_path: Path to the DOCX file

    Returns:
        Dict with extracted text
    """
    logger.info("Extracting text from DOCX", file_path=file_path)

    try:
        doc = DocxDocument(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]

        # DOCX doesn't have real pages, treat as single page
        text = "\n\n".join(paragraphs)

        return {
            "text": text,
            "page_count": 1,
            "page_texts": [text],
            "metadata": {
                "title": doc.core_properties.title or "",
                "author": doc.core_properties.author or "",
            },
        }

    except Exception as e:
        logger.error("DOCX extraction failed", error=str(e), file_path=file_path)
        raise


@app.task(bind=True, name="tasks.extract.extract_xlsx")
def extract_xlsx(self, file_path: str) -> dict:
    """Extract text from an XLSX file.

    Args:
        file_path: Path to the XLSX file

    Returns:
        Dict with extracted text (one page per sheet)
    """
    logger.info("Extracting text from XLSX", file_path=file_path)

    try:
        wb = load_workbook(file_path, data_only=True)
        page_texts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []

            for row in sheet.iter_rows():
                values = [str(cell.value) if cell.value is not None else "" for cell in row]
                if any(v.strip() for v in values):
                    rows.append("\t".join(values))

            sheet_text = f"[SHEET: {sheet_name}]\n" + "\n".join(rows)
            page_texts.append(sheet_text)

        wb.close()

        return {
            "text": "\n\n".join(page_texts),
            "page_count": len(page_texts),
            "page_texts": page_texts,
            "metadata": {"sheets": wb.sheetnames},
        }

    except Exception as e:
        logger.error("XLSX extraction failed", error=str(e), file_path=file_path)
        raise


def get_extractor(mime_type: str) -> str:
    """Get the appropriate extractor task name for a MIME type.

    Args:
        mime_type: File MIME type

    Returns:
        Task name for extraction
    """
    extractors = {
        "application/pdf": "tasks.extract.extract_pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "tasks.extract.extract_docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "tasks.extract.extract_xlsx",
        "application/msword": "tasks.extract.extract_docx",
        "application/vnd.ms-excel": "tasks.extract.extract_xlsx",
    }
    return extractors.get(mime_type, "tasks.extract.extract_pdf")
